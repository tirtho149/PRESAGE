#!/usr/bin/env python3
"""Export the PRESAGE final registry to ONE fully-visualized, color-coded Excel sheet.

Single sheet "PRESAGE_KB":
  - Title + summary stats + colour legend at the top.
  - Hierarchical, indented body: CROP banner -> DISEASE banner -> canonical fields
    -> regional deltas. Rows are grouped (outline) so crops/diseases collapse.
  - Colour-coded by delta verification_status and by link/quote liveness.
  - Frozen header, wrapped text, tuned column widths.
"""
import json, glob, os, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# Crops to pin to the very top (in this order); everything else alphabetical.
PRIORITY_CROPS = ["Soybean", "Corn", "Wheat", "Tomato"]
# Expert (plant-pathologist) verification dropdown options.
EXPERT_OPTIONS = ["Correct", "Needs revision", "Incorrect"]

KB = "artifacts/pathome_kb"
OUT = "artifacts/PRESAGE_final_registry.xlsx"
LINKCHECK = "artifacts/link_check.json"
QUOTECHECK = "artifacts/quote_check.json"

# ------------------------------------------------------------------ lookups
_lc = json.load(open(LINKCHECK)) if os.path.exists(LINKCHECK) else {}
def link_status(url):
    if not url:
        return ""
    return (_lc.get(url) or {}).get("status", "unchecked")

# path -> quote_status (did the stored quote actually appear on the cited page?)
_qc_path = {}
if os.path.exists(QUOTECHECK):
    for url, rec in json.load(open(QUOTECHECK)).items():
        for q in rec.get("quotes", []):
            if q.get("path"):
                _qc_path[q["path"]] = q.get("quote_status", "")

def join_list(v):
    if isinstance(v, list):
        return "; ".join(str(i) for i in v)
    return "" if v is None else str(v)

def s(x):
    if x is None:
        return ""
    if isinstance(x, dict):
        if "value" in x:
            return join_list(x["value"])
        if "summary" in x and isinstance(x["summary"], dict):
            return join_list(x["summary"].get("value"))
        return json.dumps(x, ensure_ascii=False)
    return join_list(x)

# ------------------------------------------------------------------ styling
def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

C_TITLE      = fill("1F3864")   # deep navy
C_CROP       = fill("2E5496")   # blue banner
C_DISEASE    = fill("8EAADB")   # light-blue banner
C_HEADER     = fill("44546A")   # table header
C_FIELD      = fill("F2F2F2")   # canonical field rows (light grey)
C_LEGEND     = fill("FFFFFF")
# verification_status colours
STATUS_FILL = {
    "verified":          fill("C6EFCE"),   # green
    "weakly_supported":  fill("E2EFDA"),   # pale green
    "field_observation": fill("DDEBF7"),   # pale blue
    "novel_plausible":   fill("FFF2CC"),   # pale yellow
    "provisional":       fill("FCE4D6"),   # pale orange
    "unverified":        fill("F8CBAD"),   # orange
    "contradictory":     fill("FFC7CE"),   # red
}
STATUS_FONT = {
    "verified":          "006100",
    "weakly_supported":  "375623",
    "field_observation": "1F4E79",
    "novel_plausible":   "7F6000",
    "provisional":       "833C00",
    "unverified":        "833C00",
    "contradictory":     "9C0006",
}
# link liveness dot colours
LINK_FILL = {"alive": fill("C6EFCE"), "blocked": fill("FFF2CC"),
             "dead": fill("FFC7CE"), "unknown": fill("E7E6E6"), "unchecked": fill("FFFFFF")}
# quote-found colours
QUOTE_FILL = {"quote_found": fill("C6EFCE"), "quote_partial": fill("FFF2CC"),
              "quote_not_found": fill("FFC7CE"), "url_blocked": fill("E7E6E6"),
              "url_dead": fill("FFC7CE"), "url_unfetched_text": fill("E7E6E6")}

WHITE_BOLD = Font(color="FFFFFF", bold=True)
HYPERLINK_FONT = Font(color="0563C1", underline="single")   # clickable-link look
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_L = Alignment(wrap_text=True, vertical="top", horizontal="left")

# ------------------------------------------------------------------ columns
# A label(indented) | B value/observation | C status | D source url | E quote
# F link | G quote-check | H notes | I EXPERT VERDICT | J EXPERT COMMENTS
COLS = ["Item", "Value / Observation", "Status", "Source URL",
        "Quote (as stored)", "Link", "Quote?", "Notes",
        "Expert Verdict", "Expert Comments"]
WIDTHS = [46, 58, 18, 50, 58, 10, 14, 44, 18, 32]
NCOL = len(COLS)
EXPERT_COL = 9          # column I
COMMENT_COL = 10        # column J
_data_runs = []         # contiguous (start_row, end_row) blocks of factual rows for the dropdown
_run = [None, None]
def _mark_data(row):
    if _run[0] is None:
        _run[0] = row
    _run[1] = row
def _close_run():
    if _run[0] is not None:
        _data_runs.append((_run[0], _run[1]))
        _run[0] = _run[1] = None

wb = Workbook()
ws = wb.active
ws.title = "PRESAGE_KB"
ws.sheet_properties.outlinePr.summaryBelow = False   # parent above its detail
r = 1

def setrow(cells, fillc=None, font=None, align=WRAP, height=None,
           indent=0, level=0, border=True):
    """Write one row of up to NCOL cells; return the row index used."""
    global r
    for i in range(NCOL):
        c = ws.cell(row=r, column=i + 1)
        c.value = cells[i] if i < len(cells) else None
        # keep the two expert columns white so the dropdown + colour rules read clearly
        if i in (EXPERT_COL - 1, COMMENT_COL - 1):
            c.fill = fill("FFFFFF")
        elif fillc:
            c.fill = fillc
        if font: c.font = font
        a = align
        if indent and i == 0:
            a = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=indent)
        c.alignment = a
        if border: c.border = BORDER
    if height: ws.row_dimensions[r].height = height
    if level: ws.row_dimensions[r].outline_level = level
    used = r
    _mark_data(used)          # every setrow() is a factual row -> gets an expert dropdown
    r += 1
    return used

def banner(text, fillc, level=0, height=20):
    global r
    _close_run()   # a banner breaks the contiguous factual-row block
    # merge only across the content columns A:H; leave the two expert columns
    # (I,J) unmerged so their dropdown / colour rules stay clean.
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL - 2)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = fillc; c.font = WHITE_BOLD
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in (EXPERT_COL, COMMENT_COL):     # tint the expert cells to match the banner
        bc = ws.cell(row=r, column=col)
        bc.fill = fillc
    ws.row_dimensions[r].height = height
    if level: ws.row_dimensions[r].outline_level = level
    r += 1

# ------------------------------------------------------------------ load data
files = sorted(glob.glob(f"{KB}/*/final_registry.json"))
status_counts = collections.Counter()
n_crops = n_dz = n_delta = n_web = 0
tree = []   # (crop, registry)
for f in files:
    d = json.load(open(f))
    crop = d.get("crop", os.path.basename(os.path.dirname(f)))
    tree.append((crop, d))
    n_crops += 1
    for dz in d.get("diseases", []):
        n_dz += 1
        for state, st in (dz.get("regional_observations") or {}).items():
            for dl in st.get("deltas", []):
                n_delta += 1
                status_counts[dl.get("verification_status", "")] += 1
                if dl.get("web_support"):
                    n_web += 1

# pin priority crops to the top (Soybean, Corn, Wheat, Tomato), rest alphabetical
def _crop_sort_key(item):
    crop = item[0]
    return (0, PRIORITY_CROPS.index(crop)) if crop in PRIORITY_CROPS else (1, crop.lower())
tree.sort(key=_crop_sort_key)

# ------------------------------------------------------------------ header block
banner("PRESAGE PathomeDB — Knowledge Base (canonical + regional deltas)", C_TITLE, height=26)
stat = (f"Crops: {n_crops}    Diseases: {n_dz}    Regional deltas: {n_delta}"
        f"    Web-cited deltas: {n_web} ({n_web*100//max(n_delta,1)}%)")
banner(stat, C_CROP, height=18)
# legend row
r += 0
ws.cell(row=r, column=1, value="Legend (delta status):").font = Font(bold=True)
legend = list(STATUS_FILL.items())
for i, (name, fl) in enumerate(legend):
    c = ws.cell(row=r, column=2 + i, value=name)
    c.fill = fl
    c.font = Font(color=STATUS_FONT.get(name, "000000"), bold=True, size=9)
    c.alignment = Alignment(horizontal="center")
ws.row_dimensions[r].height = 16
r += 1
ws.cell(row=r, column=1, value="Legend (link / quote):").font = Font(bold=True)
for i, (name, fl) in enumerate([("alive", LINK_FILL["alive"]), ("blocked", LINK_FILL["blocked"]),
                                ("dead", LINK_FILL["dead"]), ("quote_found", QUOTE_FILL["quote_found"]),
                                ("quote_not_found", QUOTE_FILL["quote_not_found"])]):
    c = ws.cell(row=r, column=2 + i, value=name)
    c.fill = fl; c.font = Font(bold=True, size=9); c.alignment = Alignment(horizontal="center")
ws.row_dimensions[r].height = 16
r += 1
r += 1  # spacer

# ------------------------------------------------------------------ table header
hdr_row = r
for i, name in enumerate(COLS):
    c = ws.cell(row=r, column=i + 1, value=name)
    c.fill = C_HEADER; c.font = WHITE_BOLD
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws.row_dimensions[r].height = 22
r += 1
ws.freeze_panes = ws.cell(row=r, column=1)   # freeze everything above the first data row

# ------------------------------------------------------------------ body
CANON_FIELDS = [
    ("Pathogen", "pathogen_scientific_name"),
    ("Disease type", "type_of_disease"),
    ("Affected parts", "affected_parts"),
    ("Treatments", "treatments"),
]

def qstatus(path):
    return _qc_path.get(path, "")

URL_COL = 4   # column D "Source URL"
def set_link(row_idx, url):
    """Make the Source URL cell a real clickable hyperlink to the page."""
    if not url:
        return
    c = ws.cell(row=row_idx, column=URL_COL)
    c.hyperlink = url
    c.value = url
    c.font = HYPERLINK_FONT
    c.alignment = WRAP_L

for crop, d in tree:
    dzs = d.get("diseases", [])
    ndelt = sum(len(st.get("deltas", []))
                for dz in dzs for st in (dz.get("regional_observations") or {}).values())
    banner(f"🌱  {crop}    —    {len(dzs)} disease(s), {ndelt} regional delta(s)", C_CROP, level=0, height=22)
    for di, dz in enumerate(dzs):
        name = dz.get("disease_name", "")
        pth = s(dz.get("pathogen_scientific_name"))
        typ = s(dz.get("type_of_disease"))
        banner(f"      🦠  {name}    ·    {pth}    ·    {typ}", C_DISEASE, level=1, height=18)
        base = f"{crop}/diseases[{di}]"
        # canonical fields
        for label, key in CANON_FIELDS:
            node = dz.get(key)
            if node is None:
                continue
            val = s(node)
            url = node.get("url", "") if isinstance(node, dict) else ""
            quote = node.get("quote", "") if isinstance(node, dict) else ""
            ls = link_status(url)
            qs = qstatus(f"{base}/{key}")
            row = setrow([f"• {label}", val, "", url, quote, ls, qs, ""],
                         fillc=C_FIELD, indent=2, level=2)
            set_link(row, url)
            if ls: ws.cell(row=row, column=6).fill = LINK_FILL.get(ls, C_FIELD)
            if qs: ws.cell(row=row, column=7).fill = QUOTE_FILL.get(qs, C_FIELD)
        # visual symptoms (summary + diagnostic_features live nested)
        vs = dz.get("visual_symptoms") or {}
        for sub, slabel in [("summary", "Visual symptoms"), ("diagnostic_features", "Diagnostic features")]:
            node = vs.get(sub)
            if isinstance(node, dict) and (node.get("value") or node.get("quote")):
                url = node.get("url", ""); quote = node.get("quote", "")
                ls = link_status(url); qs = qstatus(f"{base}/visual_symptoms/{sub}")
                row = setrow([f"• {slabel}", join_list(node.get("value")), "", url, quote, ls, qs, ""],
                             fillc=C_FIELD, indent=2, level=2)
                set_link(row, url)
                if ls: ws.cell(row=row, column=6).fill = LINK_FILL.get(ls, C_FIELD)
                if qs: ws.cell(row=row, column=7).fill = QUOTE_FILL.get(qs, C_FIELD)
        # regional deltas
        for state, st in (dz.get("regional_observations") or {}).items():
            stname = st.get("state", state)
            for li, dl in enumerate(st.get("deltas", [])):
                status = dl.get("verification_status", "")
                ws_entries = [w for w in (dl.get("web_support") or []) if w.get("url")]
                nsrc = len(ws_entries)
                notes = []
                if dl.get("_prior_status"): notes.append(f"was: {dl['_prior_status']}")
                if dl.get("image_id"): notes.append(dl["image_id"])
                if dl.get("reasoning"): notes.append(dl["reasoning"])
                fld = dl.get("field", "")
                src_summary = (f"{nsrc} source(s) ↓" if nsrc
                               else "— no web source (field observation)")
                row = setrow([f"Δ  {stname} · {fld}",
                              dl.get("image_shows", ""), status, src_summary,
                              dl.get("image_quote", ""), "", "",
                              "  |  ".join(notes)],
                             fillc=STATUS_FILL.get(status, None), indent=3, level=2)
                # colour the status cell strongly + status font
                sc = ws.cell(row=row, column=3)
                sc.fill = STATUS_FILL.get(status, C_FIELD)
                sc.font = Font(color=STATUS_FONT.get(status, "000000"), bold=True)
                sc.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                # one clickable sub-row per web source (so EVERY link is tappable)
                base = f"{crop}/diseases[{di}]/regional_observations/{state}/deltas[{li}]"
                for wi, w in enumerate(ws_entries):
                    u = w.get("url", ""); q = w.get("quote", "")
                    lsw = link_status(u); qs = qstatus(f"{base}/web_support[{wi}]")
                    srow = setrow([f"      ↳ source {wi+1}", "", "", u, q, lsw, qs, ""],
                                  fillc=C_FIELD, indent=4, level=3)
                    set_link(srow, u)
                    if lsw: ws.cell(row=srow, column=6).fill = LINK_FILL.get(lsw, C_FIELD)
                    if qs: ws.cell(row=srow, column=7).fill = QUOTE_FILL.get(qs, C_FIELD)

# ------------------------------------------------------------------ expert verification
_close_run()   # close the final factual-row block
col = get_column_letter(EXPERT_COL)   # "I"
# 3-option plant-pathologist dropdown on every factual row (canonical + delta)
dv = DataValidation(type="list", formula1='"%s"' % ",".join(EXPERT_OPTIONS),
                    allow_blank=True, showErrorMessage=True)
dv.promptTitle = "Expert verification"
dv.prompt = "Plant pathologist: is this entry correct?"
dv.errorTitle = "Pick one"
dv.error = "Choose: " + " / ".join(EXPERT_OPTIONS)
ws.add_data_validation(dv)
n_cells = 0
for a, b in _data_runs:
    dv.add(f"{col}{a}:{col}{b}")
    n_cells += b - a + 1
# colour the verdict cell by the expert's choice
if _data_runs:
    lo = _data_runs[0][0]; hi = _data_runs[-1][1]
    rng = f"{col}{lo}:{col}{hi}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Correct"'],
        fill=fill("C6EFCE"), font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Needs revision"'],
        fill=fill("FFF2CC"), font=Font(color="7F6000", bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Incorrect"'],
        fill=fill("FFC7CE"), font=Font(color="9C0006", bold=True)))

# ------------------------------------------------------------------ widths & finish
for i, w in enumerate(WIDTHS):
    ws.column_dimensions[get_column_letter(i + 1)].width = w
ws.sheet_view.showGridLines = False

wb.save(OUT)
print(f"  expert-verification dropdowns on {n_cells} factual rows ({len(_data_runs)} blocks)")
print(f"WROTE {OUT}")
print(f"  crops={n_crops}  diseases={n_dz}  deltas={n_delta}  web_cited={n_web}")
print(f"  status={dict(status_counts)}")
print(f"  rows written={r-1}")
