"""Generate disease_registry_validation_v2.xlsx — multi-crop expert
validation panel, formatted in the style of the original
disease_registry_validation.xlsx (marker + disease-name-stripe + data
row pattern per canonical field, closing disease-name row + spacer per
disease). At the bottom of each disease block — before the closing row
— a self-contained "Regional Deltas by State" section lists every state
that has image-grounded observations, with that state's full delta KB
in col D and its own Delta Verdict dropdown.

Crops sorted by disease count (most diseases at the top), built from
artifacts/pathome_kb/<Crop>/final_registry.json across all 14 crops.

Layout (10 cols):
  A  Disease  /  Disease · State
  B  Field  /  State
  C  Canonical Claim / Image Observation
  D  Delta KB (image-grounded; populated on per-state rows)
  E  Source URL
  F  Quote #  /  Status
  G  Full Quote / Evidence
  H  Canonical Verdict   (dropdown on canonical rows; "—" elsewhere)
  I  Delta Verdict       (dropdown on per-state rows; "—" elsewhere)
  J  Validator Notes
"""
from __future__ import annotations

import copy
import io
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

KB_ROOT  = Path("artifacts/pathome_kb")
TEMPLATE = Path("artifacts/phase0r_traces/disease_registry_validation.xlsx")
OUT      = Path("artifacts/phase0r_traces/disease_registry_validation_v2.xlsx")

NUM_COLS = 10

# Source landmark rows in the template (Soybean / Anthracnose block)
ROW_TITLE         = 1
ROW_INSTRUCT      = 2
ROW_HEADER        = 3
ROW_CROP_BANNER   = 4
ROW_CROP_INSTRUCT = 5
ROW_DIS_BANNER    = 6
ROW_FIELD_MARKER  = 7
ROW_DIS_NAME      = 8
ROW_DATA          = 9
ROW_LOOKALIKES    = 23
ROW_DIS_CLOSE     = 24
ROW_SPACER        = 25
ROW_CONFLICT      = 27

CROP_EMOJI = {
    "Soybean": "🫘", "Wheat": "🌾", "Corn": "🌽", "Apple": "🍎",
    "Watermelon": "🍉", "Melon": "🍈", "Tomato": "🍅", "Rice": "🌾",
    "Sweet Potato": "🍠", "Oat": "🌾", "Alder": "🌳", "Spruce": "🌲",
    "Mallow": "🌸", "Kentucky Bluegrass": "🌿",
}

FIELDS = [
    ("visual_symptoms.summary",             "Visual Symptoms – Summary"),
    ("visual_symptoms.diagnostic_features", "Visual Symptoms – Diagnostic Features"),
    ("visual_symptoms.look_alikes",         "Visual Symptoms – Look-alikes"),
]

STATUS_BADGE = {
    "verified":         "✓ verified",
    "weakly_supported": "~ weakly_supported",
    "provisional":      "? provisional",
    "novel_plausible":  "* novel_plausible",
    "unverified":       "! unverified",
}
STATUS_RANK = {
    "verified": 0, "weakly_supported": 1, "provisional": 2,
    "novel_plausible": 3, "unverified": 4,
}


# ---------------------------------------------------------------------------
# Style capture — extend each template row to 10 cells:
#   insert a clone of col-3 (canonical claim) at index 4 → "Delta KB"
#   insert a clone of col-7 (Verdict) at index 9 → "Delta Verdict"
# Original 8 cells [c1..c8] -> new 10 cells:
#   [c1, c2, c3, c3_clone, c4, c5, c6, c7, c7_clone, c8]
# ---------------------------------------------------------------------------

def _capture(src_ws, src_row: int) -> dict:
    src = [_capture_cell(src_ws.cell(src_row, c)) for c in range(1, 9)]
    cells = [
        src[0], src[1], src[2],
        copy.deepcopy(src[2]),   # D = Delta KB (clone of C)
        src[3], src[4], src[5], src[6],
        copy.deepcopy(src[6]),   # I = Delta Verdict (clone of Verdict)
        src[7],                  # J = Validator Notes
    ]
    return {"height": src_ws.row_dimensions[src_row].height, "cells": cells}


def _capture_cell(cell) -> dict:
    return {
        "font":      copy.copy(cell.font),
        "fill":      copy.copy(cell.fill),
        "border":    copy.copy(cell.border),
        "alignment": copy.copy(cell.alignment),
    }


def _apply(dst_ws, dst_row: int, template: dict, values=None,
           merge_full_width: bool = False):
    dst_ws.row_dimensions[dst_row].height = template["height"]
    for c in range(1, NUM_COLS + 1):
        cell = dst_ws.cell(dst_row, c)
        cell.value = (values[c-1] if values and c-1 < len(values) else None)
        style = template["cells"][c-1]
        cell.font      = copy.copy(style["font"])
        cell.fill      = copy.copy(style["fill"])
        cell.border    = copy.copy(style["border"])
        cell.alignment = copy.copy(style["alignment"])
    if merge_full_width:
        dst_ws.merge_cells(start_row=dst_row, start_column=1,
                           end_row=dst_row, end_column=NUM_COLS)


# ---------------------------------------------------------------------------
# Field accessors
# ---------------------------------------------------------------------------

def _resolve(disease: dict, dotted: str) -> dict | None:
    cur = disease
    for part in dotted.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(part)
    return cur if isinstance(cur, dict) else None


def _fmt_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v if x)
    return str(v)


def _delta_text(delta: dict) -> str:
    """Format one delta as a single-line ``[status] field: image_shows``."""
    status = delta.get("verification_status") or "unverified"
    field  = delta.get("field") or "other"
    shows  = (delta.get("image_shows") or "").strip()
    return f"[{STATUS_BADGE.get(status, status)}] {field}: {shows}"


def _delta_url_quote_pairs(delta: dict) -> list[tuple[str, str]]:
    """Unique (url, quote) pairs for a single delta's ``web_support``.
    Direct URLs first; inherited URLs tagged ``(inherited)``."""
    direct: dict[str, str] = {}
    inherited: dict[str, str] = {}
    for w in delta.get("web_support") or []:
        if not isinstance(w, dict):
            continue
        url = (w.get("url") or "").strip()
        if not url:
            continue
        quote = (w.get("quote") or "").strip()
        if w.get("inherited"):
            if url not in direct and url not in inherited:
                inherited[url] = quote
        else:
            if url not in direct:
                direct[url] = quote
            inherited.pop(url, None)
    pairs: list[tuple[str, str]] = [(u, q) for u, q in direct.items()]
    pairs += [(f"{u}  (inherited)", q) for u, q in inherited.items()]
    return pairs


# ---------------------------------------------------------------------------
# Disease block writer
# ---------------------------------------------------------------------------

def write_disease(ws, row: int, disease: dict, templates: dict,
                  canonical_verdict_rows: list,
                  delta_verdict_rows: list) -> int:
    name = disease.get("disease_name", "")
    dtype_field = disease.get("type_of_disease") or {}
    dtype = dtype_field.get("value") if isinstance(dtype_field, dict) else dtype_field
    confidence = (disease.get("confidence") or "—").capitalize()
    num_sources = disease.get("num_sources", "—")
    conflicts = disease.get("conflicts") or []

    # Disease banner (merged A:J)
    header = (f"▶  Disease: {name}   |   Type: {dtype or '—'}   |   "
              f"Confidence: {confidence}   |   Sources: {num_sources}")
    _apply(ws, row, templates["dis_banner"],
           [header] + [None]*(NUM_COLS-1), merge_full_width=True)
    row += 1

    if conflicts:
        _apply(ws, row, templates["conflict"],
               [f"   ⚠️  {len(conflicts)} conflict(s) recorded"] + [None]*(NUM_COLS-1),
               merge_full_width=True)
        row += 1

    # ------------------------------------------------------------------
    # Canonical field blocks (original 3-row pattern: marker, stripe, data;
    # look_alikes uses the condensed 2-row pattern).
    # ------------------------------------------------------------------
    for dotted, pretty in FIELDS:
        field_obj = _resolve(disease, dotted)
        value = _fmt_value(field_obj.get("value")) if field_obj else ""
        url   = (field_obj.get("url") if field_obj else "") or ""
        quote = (field_obj.get("quote") if field_obj else "") or ""

        is_look_alikes = dotted.endswith("look_alikes")

        # Field marker row
        _apply(ws, row, templates["field_marker"],
               [f"   Field: {dotted}"] + [None]*(NUM_COLS-1))
        row += 1

        if is_look_alikes:
            _apply(ws, row, templates["look_alikes"],
                   [None, pretty, value, "—", url, "—", None, None, "—", None])
            canonical_verdict_rows.append(row)
            row += 1
        else:
            _apply(ws, row, templates["dis_name"],
                   [name] + [None]*(NUM_COLS-1))
            row += 1
            quote_tag = "Quote 1 of 1 (full text)" if quote else "—"
            _apply(ws, row, templates["data"],
                   [None, pretty, value, "—", url, quote_tag, quote,
                    None, "—", None])
            canonical_verdict_rows.append(row)
            row += 1

    # ------------------------------------------------------------------
    # Regional Deltas by State — bottom section.
    # One row per state that has image-grounded observations. Col D
    # carries every delta for that state (full text, status badges).
    # Each row gets its own Delta Verdict dropdown in col I.
    # ------------------------------------------------------------------
    ro = disease.get("regional_observations") or {}
    states_with_deltas = sorted(
        s for s in ro if (ro.get(s) or {}).get("deltas")
    )
    if states_with_deltas:
        n_deltas = sum(len((ro[s] or {}).get("deltas") or [])
                       for s in states_with_deltas)
        _apply(ws, row, templates["field_marker"],
               [f"   Regional Deltas by State  ({len(states_with_deltas)} "
                f"state(s), {n_deltas} image-grounded deltas)"]
               + [None]*(NUM_COLS-1))
        row += 1

        for state in states_with_deltas:
            deltas = (ro[state] or {}).get("deltas") or []
            # One row per (delta, source) pair. Each delta's text in col D
            # is REPEATED across every source row for that delta so col D
            # always shows the claim the URL/quote in cols E-G supports.
            # State header (A,B,C) appears on the very first row of the
            # state only; the Delta Verdict dropdown appears on the first
            # row of each delta only.
            state_first_row = True
            for delta in sorted(
                deltas,
                key=lambda d: STATUS_RANK.get(
                    d.get("verification_status") or "unverified", 99
                ),
            ):
                d_text = _delta_text(delta)
                pairs = _delta_url_quote_pairs(delta)
                n = len(pairs)
                if n == 0:
                    pairs = [(None, None)]
                    n = 0  # signals "no sources" for the Quote tag below
                for i, (url, quote) in enumerate(pairs, start=1):
                    is_first_of_delta  = (i == 1)
                    is_first_of_state  = state_first_row and is_first_of_delta
                    tag = (f"Quote {i} of {n}") if n else (
                        "(no sources)" if is_first_of_delta else None
                    )
                    _apply(ws, row, templates["data"], [
                        f"{name} · {state}" if is_first_of_state else None,
                        state                if is_first_of_state else None,
                        (f"{len(deltas)} image-grounded observation(s)"
                         if is_first_of_state else None),
                        d_text if is_first_of_delta else None,
                        url,
                        tag,
                        quote,
                        "—",  # canonical verdict N/A
                        None if is_first_of_delta else "—",
                        None,
                    ])
                    if is_first_of_delta:
                        delta_verdict_rows.append(row)
                    row += 1
                    state_first_row = False

    # Closing disease-name row + spacer
    _apply(ws, row, templates["dis_close"], [name] + [None]*(NUM_COLS-1))
    row += 1
    _apply(ws, row, templates["spacer"], [None]*NUM_COLS)
    row += 1
    return row


# ---------------------------------------------------------------------------
# Crop block writer
# ---------------------------------------------------------------------------

def write_crop(ws, row: int, crop: str, diseases: list, templates: dict,
               canonical_verdict_rows: list,
               delta_verdict_rows: list) -> int:
    emoji = CROP_EMOJI.get(crop, "🌱")
    today = date.today().isoformat()
    banner = (f"{emoji}  CROP: {crop.upper()}  —  Registry  |  "
              f"{len(diseases)} diseases  |  Generated: {today}")
    _apply(ws, row, templates["crop_banner"],
           [banner] + [None]*(NUM_COLS-1), merge_full_width=True)
    row += 1
    _apply(ws, row, templates["crop_instruct"],
           ["  Review each canonical claim against the quoted source text. "
            "Then validate the per-state image-grounded observations in the "
            "Regional Deltas by State section. Canonical Verdict (col H) for "
            "canonical-claim rows; Delta Verdict (col I) for state rows."]
           + [None]*(NUM_COLS-1), merge_full_width=True)
    row += 1
    for d in sorted(diseases, key=lambda x: (x.get("disease_name") or "").lower()):
        row = write_disease(ws, row, d, templates,
                            canonical_verdict_rows, delta_verdict_rows)
    return row


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def main():
    src_wb = load_workbook(TEMPLATE, data_only=False)
    src_ws = src_wb["Validation Panel"]
    templates = {
        "title":          _capture(src_ws, ROW_TITLE),
        "instruct":       _capture(src_ws, ROW_INSTRUCT),
        "header":         _capture(src_ws, ROW_HEADER),
        "crop_banner":    _capture(src_ws, ROW_CROP_BANNER),
        "crop_instruct":  _capture(src_ws, ROW_CROP_INSTRUCT),
        "dis_banner":     _capture(src_ws, ROW_DIS_BANNER),
        "field_marker":   _capture(src_ws, ROW_FIELD_MARKER),
        "dis_name":       _capture(src_ws, ROW_DIS_NAME),
        "data":           _capture(src_ws, ROW_DATA),
        "look_alikes":    _capture(src_ws, ROW_LOOKALIKES),
        "dis_close":      _capture(src_ws, ROW_DIS_CLOSE),
        "spacer":         _capture(src_ws, ROW_SPACER),
        "conflict":       _capture(src_ws, ROW_CONFLICT),
    }
    src_widths = {letter: dim.width
                  for letter, dim in src_ws.column_dimensions.items()
                  if dim.width}

    crops = []
    for cd in sorted(p for p in KB_ROOT.iterdir() if p.is_dir()):
        rp = cd / "final_registry.json"
        if not rp.is_file():
            continue
        reg = json.loads(rp.read_text(encoding="utf-8"))
        diseases = reg.get("diseases") or []
        if not diseases:
            continue
        crops.append((cd.name, diseases))
    # Soybean and Corn pinned to the top; remaining crops by disease
    # count (descending), then alphabetical.
    _PINNED = {"Soybean": 0, "Corn": 1}
    crops.sort(key=lambda x: (_PINNED.get(x[0], 2), -len(x[1]), x[0]))

    print(f"crops ({len(crops)}), ordered by disease count:")
    for c, dl in crops:
        print(f"  {c:<22} {len(dl)} diseases")

    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Panel"

    widths = {
        "A": src_widths.get("A", 43.7),
        "B": src_widths.get("B", 30.0),
        "C": src_widths.get("C", 32.0),
        "D": 70.0,                                  # Delta KB column
        "E": src_widths.get("D", 36.0),             # Source URL
        "F": max(src_widths.get("E", 9.0), 22.0),   # Quote # / Status
        "G": src_widths.get("F", 58.0),             # Full Quote / Evidence
        "H": src_widths.get("G", 18.0),             # Canonical Verdict
        "I": src_widths.get("G", 18.0),             # Delta Verdict
        "J": src_widths.get("H", 30.0),             # Validator Notes
    }
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    _apply(ws, 1, templates["title"],
           ["🔬  EXPERT VALIDATION PANEL — MULTI-CROP DISEASE REGISTRY"] + [None]*(NUM_COLS-1),
           merge_full_width=True)
    _apply(ws, 2, templates["instruct"],
           ["Instructions:  Review each canonical claim against the quoted "
            "source. After the 6 canonical fields, each disease ends with a "
            "Regional Deltas by State section — one row per state with all "
            "image-grounded observations for that state in col D. Use "
            "Canonical Verdict (H) for canonical rows and Delta Verdict (I) "
            "for state rows. Cells marked '—' indicate that verdict doesn't "
            "apply to the row."] + [None]*(NUM_COLS-1),
           merge_full_width=True)
    _apply(ws, 3, templates["header"],
           ["Disease  /  Disease · State", "Field  /  State",
            "Canonical Claim  /  Image Observation",
            "Delta KB (image-grounded)",
            "Source URL", "Quote #  /  Status",
            "Full Quote  /  Evidence",
            "Canonical Verdict", "Delta Verdict", "Validator Notes"])

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:J3"

    canonical_verdict_rows: list[int] = []
    delta_verdict_rows: list[int] = []
    row = 4
    for crop, diseases in crops:
        row = write_crop(ws, row, crop, diseases, templates,
                         canonical_verdict_rows, delta_verdict_rows)

    # Chunked dropdowns — Excel drops a DataValidation whose sqref
    # attribute exceeds ~8000 chars (cell refs are ~5 chars apiece, so
    # 250 cells per chunk keeps it well under).
    CHUNK = 250

    def _mk_dv(title: str, prompt: str) -> DataValidation:
        return DataValidation(
            type="list",
            formula1='"Agree,Neutral,Disagree"',
            allow_blank=True,
            showDropDown=False,
            errorStyle="warning",
            errorTitle="Invalid verdict",
            error="Pick one of Agree / Neutral / Disagree.",
            promptTitle=title,
            prompt=prompt,
            showErrorMessage=True,
            showInputMessage=False,
        )

    def _attach_chunked(rows: list[int], col: str, title: str, prompt: str) -> int:
        n = 0
        for i in range(0, len(rows), CHUNK):
            dv = _mk_dv(title, prompt)
            for r in rows[i:i+CHUNK]:
                dv.add(f"{col}{r}")
            ws.add_data_validation(dv)
            n += 1
        return n

    n_cv = _attach_chunked(canonical_verdict_rows, "H",
                           "Canonical Verdict",
                           "Agree / Neutral / Disagree with the canonical claim.")
    n_dv = _attach_chunked(delta_verdict_rows, "I",
                           "Delta Verdict",
                           "Agree / Neutral / Disagree with the state's image-grounded delta KB.")
    print(f"  Canonical Verdict dropdown: {len(canonical_verdict_rows)} cells "
          f"(col H, {n_cv} chunks)")
    print(f"  Delta     Verdict dropdown: {len(delta_verdict_rows)} cells "
          f"(col I, {n_dv} chunks)")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"\nwrote {OUT}  ({row-1} rows)")


if __name__ == "__main__":
    main()
